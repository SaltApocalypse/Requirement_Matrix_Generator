from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment, Border, Side

import utils.math
from utils.Sheettree import SheetTree


def is_merged_cell(ws: Workbook, cell: str) -> bool:
    """
    检查某个单元格是否是合并单元格。

    Args:
        ws (Workbook): 包含工作表的工作簿。
        cell (str): 需要检查的具体单元格，比如 `A1` 。

    Returns:
        bool: 如果指定单元格在合并单元格内则返回 `True` 否则返回 `False` 。
    """
    cell = ws[cell]
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        cell_col = cell.column
        cell_row = cell.row
        if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
            return True
    return False


def create_workbook(column: SheetTree, row: SheetTree, title: str, path: str):
    """
    从结构树生成需求矩阵，并写入新建的工作簿。

    Args:
        column (SheetTree): 处理好的列结构。
        row (SheetTree): 处理好的行结构。
        title (str): 矩阵的标题。
        path (str): 保存的路径。

    Returns:
        Workbook: 包含需求矩阵的新建工作簿对象。
    """
    workbook = Workbook()
    worksheet = workbook.active

    column_length = column.tree_depth
    row_height = row.tree_depth

    # title
    worksheet["A1"] = title
    worksheet.merge_cells("A1:" + utils.math.number_to_alpha(column_length) + str(row_height))

    # ========== 创建行标题 ==========
    sheet_number = row_height
    for i in range(column.node_nums):
        sheet_alpha = utils.math.number_to_alpha(column.node_level[i], 1)
        if 0 == i or column.node_level[i] <= column.node_level[i - 1]:
            sheet_number += 1
        sheet = sheet_alpha + str(sheet_number)
        worksheet[sheet] = column.tree_data[i]
    column_height = sheet_number - row_height

    # 合并空白单元格：每行从右往左检查
    for i in range(column_height):
        tail_sheet_alpha = utils.math.number_to_alpha(column_length)
        tail_sheet_number = column_length + i
        tail_sheet = tail_sheet_alpha + str(tail_sheet_number)

        if None == worksheet[tail_sheet].value:
            for j in range(1, column_length):
                head_sheet_alpha = utils.math.number_to_alpha(column_length, -j)
                head_sheet = head_sheet_alpha + str(tail_sheet_number)
                if worksheet[head_sheet].value != None:
                    worksheet.merge_cells(head_sheet + ":" + tail_sheet)
                    break

    # 合并空白单元格：每列从上往下检查
    for i in range(1, column_length):
        head_sheet_alpha = utils.math.number_to_alpha(i)
        head_sheet_number = tail_sheet_number = row_height + 1
        head_sheet = head_sheet_alpha + str(head_sheet_number)
        while head_sheet_number <= row_height + column_height:
            tail_sheet_number += 1
            tail_sheet = head_sheet_alpha + str(tail_sheet_number)
            if worksheet[tail_sheet].value != None or tail_sheet_number > row_height + column_height:
                tail_sheet = head_sheet_alpha + str(tail_sheet_number - 1)
                if head_sheet_number != tail_sheet_number and (not is_merged_cell(worksheet, tail_sheet)):
                    worksheet.merge_cells(head_sheet + ":" + tail_sheet)
                head_sheet_number = tail_sheet_number
                head_sheet = head_sheet_alpha + str(head_sheet_number)

    # ========== 创建列标题 ==========
    sheet_alpha = utils.math.number_to_alpha(column_length)
    for i in range(row.node_nums):
        sheet_number = row.node_level[i] + 1
        if 0 == i or row.node_level[i] <= row.node_level[i - 1]:
            sheet_alpha = utils.math.number_to_alpha(utils.math.alpha_to_number(sheet_alpha), 1)
        sheet = sheet_alpha + str(sheet_number)
        worksheet[sheet] = row.tree_data[i]
    row_length = utils.math.alpha_to_number(sheet_alpha) - column_length

    # 合并空白单元格：每行自底向上检查
    for i in range(1, row_length):
        tail_sheet_alpha = utils.math.number_to_alpha(column_length, i)
        tail_sheet_number = row_height
        tail_sheet = tail_sheet_alpha + str(tail_sheet_number)

        if None == worksheet[tail_sheet].value:
            for j in range(1, row_height):
                head_sheet_number = row_height - j
                head_sheet = tail_sheet_alpha + str(head_sheet_number)
                if worksheet[head_sheet].value != None:
                    worksheet.merge_cells(head_sheet + ":" + tail_sheet)
                    break

    # 合并空白单元格：每行从左往右检查
    for i in range(1, row_height):
        head_sheet_alpha = tail_sheet_alpha = utils.math.number_to_alpha(column_length, 1)
        head_sheet_number = i
        head_sheet = head_sheet_alpha + str(head_sheet_number)
        while utils.math.alpha_to_number(head_sheet_alpha) <= column_length + row_length:
            tail_sheet_alpha = utils.math.number_to_alpha(utils.math.alpha_to_number(tail_sheet_alpha), 1)
            tail_sheet = tail_sheet_alpha + str(head_sheet_number)
            if worksheet[tail_sheet].value != None or utils.math.alpha_to_number(tail_sheet_alpha) > column_length + row_length:
                tail_sheet = utils.math.number_to_alpha(utils.math.alpha_to_number(tail_sheet_alpha), -1) + str(head_sheet_number)
                if head_sheet_alpha != tail_sheet_alpha and (not is_merged_cell(worksheet, tail_sheet)):
                    worksheet.merge_cells(head_sheet + ":" + tail_sheet)
                head_sheet_alpha = tail_sheet_alpha
                head_sheet = head_sheet_alpha + str(head_sheet_number)

    # ========== 调整格式 ==========
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    for row in worksheet.iter_rows(min_row=1, max_row=column_height + row_height, min_col=1, max_col=column_length + row_length):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    workbook.save(path)


def generate_requirement_matrix(column: SheetTree, row: SheetTree, **kwargs):
    """
    生成需求矩阵。

    Args:
        column (SheetTree): 处理好的列结构。
        row (SheetTree): 处理好的行结构。
        kwargs: 额外的参数：output_path, name, title，来自用于传入 args.
    """
    # 生成路径
    path, name, title = kwargs["output_path"], kwargs["name"], kwargs["title"]
    if not name.endswith(".xlsx"):
        name += ".xlsx"
    path = Path(path) / name
    full_path = path.resolve()

    # 生成文件
    create_workbook(column, row, title, full_path)
